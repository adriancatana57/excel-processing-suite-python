# Advanced_VLookup_App
# Copyright (c) 2025 Adrian CATANA
# Licensed under CC BY-NC 4.0 — https://creativecommons.org/licenses/by-nc/4.0/
# https://github.com/[adriancatana57]/[excel-processing-suite-python/excel_suite.py]
import os
import sys
import csv
import datetime
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk

# --- CustomTkinter Configuration ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# === Busy dialog (with custom message) ===
def _center_window(win, parent):
    win.update_idletasks()
    px, py = parent.winfo_rootx(), parent.winfo_rooty()
    pw, ph = parent.winfo_width(), parent.winfo_height()
    ww, wh = win.winfo_width(), win.winfo_height()
    x = px + max(0, (pw - ww) // 2)
    y = py + max(0, (ph - wh) // 2)
    win.geometry(f"+{x}+{y}")


def run_with_busy(parent, func, callback=None, *args, message=None, **kwargs):
    dlg = ctk.CTkToplevel(parent)
    dlg.title("Processing…")
    dlg.transient(parent)
    dlg.resizable(False, False)
    dlg.grab_set()

    frm = ctk.CTkFrame(dlg, fg_color="transparent")
    frm.pack(fill="both", expand=True, padx=20, pady=20)

    ctk.CTkLabel(frm, text=(message or "Processing data… please wait."), font=ctk.CTkFont(weight="bold")).pack(pady=(0, 10))

    pb = ctk.CTkProgressBar(frm, mode="indeterminate", width=240)
    pb.pack()
    pb.start()

    def _set_enabled(widget, enabled: bool):
        for child in widget.winfo_children():
            try:
                if isinstance(child, ctk.CTkScrollableFrame) or isinstance(child, ctk.CTkFrame):
                    _set_enabled(child, enabled)
                else:
                    child.configure(state=("normal" if enabled else "disabled"))
            except Exception:
                pass

    try:
        parent.configure(cursor="wait")
    except Exception:
        pass
    _set_enabled(parent, False)
    parent.update_idletasks()
    _center_window(dlg, parent)

    def worker():
        result = None
        exc = None
        try:
            result = func(*args, **kwargs)
        except Exception as e:
            exc = e
        finally:
            def finalize():
                try:
                    pb.stop()
                    dlg.grab_release()
                    dlg.destroy()
                except Exception:
                    pass
                try:
                    parent.configure(cursor="")
                except Exception:
                    pass
                _set_enabled(parent, True)
                parent.update_idletasks()
                if callback:
                    callback(result, exc)
                elif exc is not None:
                    messagebox.showerror("Error", str(exc), parent=parent)
            parent.after(0, finalize)

    threading.Thread(target=worker, daemon=True).start()


# DPI awareness (Windows)
try:
    import ctypes  # noqa
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        pass
except Exception:
    pass

APP_NAME = "Advanced VLookup Application"

try:
    import chardet
except Exception:
    chardet = None

try:
    import pandas as pd  # noqa
except Exception:
    pd = None


# ---------- I/O Utilities ----------

TEXT_SEP_DEFAULT_TXT = ";"
TEXT_SEP_DEFAULT_CSV = ","

def detect_encoding(file_path, sample_size=200_000):
    if chardet is None:
        return "utf-8-sig"
    try:
        with open(file_path, "rb") as f:
            raw = f.read(sample_size)
        result = chardet.detect(raw)
        enc = result.get("encoding") or "utf-8-sig"
        return enc
    except Exception:
        return "utf-8-sig"

def detect_separator(file_path, encoding=None, sample_size=200_000):
    if encoding is None:
        encoding = detect_encoding(file_path)
    try:
        with open(file_path, "r", encoding=encoding, errors="replace", newline="") as f:
            sample = f.read(sample_size)
        if not sample:
            return ","
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        return dialect.delimiter
    except Exception:
        for sep in [",", ";", "\t", "|"]:
            try:
                with open(file_path, "r", encoding=encoding, errors="replace", newline="") as f:
                    first_line = f.readline()
                if sep in first_line:
                    return sep
            except Exception:
                pass
        return ","

def require_openpyxl():
    try:
        import openpyxl as op
        return op
    except Exception as e:
        raise RuntimeError("Reading .xlsx files requires 'openpyxl' (not installed or failed to load).") from e

def require_xlsxwriter():
    try:
        import xlsxwriter as xw
        return xw
    except Exception as e:
        raise RuntimeError("Saving .xlsx files requires the 'xlsxwriter' package.") from e

def read_text_header(file_path, sep=None, encoding=None):
    if encoding is None:
        encoding = detect_encoding(file_path)
    if sep is None:
        sep = detect_separator(file_path, encoding=encoding)
    with open(file_path, "r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=sep)
        try:
            headers = next(reader, None)
        except Exception:
            headers = None
    if headers is None:
        headers = []
    headers = [str(h) if h is not None else "" for h in headers]
    seen = {}
    unique_headers = []
    for h in headers:
        if h not in seen:
            seen[h] = 0
            unique_headers.append(h)
        else:
            seen[h] += 1
            unique_headers.append(f"{h}.{seen[h]}")
    return unique_headers, sep, encoding

def iter_text_rows(file_path, sep, encoding, skip_header=True):
    with open(file_path, "r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=sep)
        if skip_header:
            next(reader, None)
        for row in reader:
            yield ["" if x is None else str(x) for x in row]

def get_xlsx_sheet_names(file_path):
    op = require_openpyxl()
    wb = op.load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames[:]
    wb.close()
    return names

def read_xlsx_header(file_path, sheet_name=None):
    op = require_openpyxl()
    wb = op.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = ["" if c is None else str(c) for c in row]
        break
    wb.close()
    seen = {}
    unique_headers = []
    for h in headers:
        if h not in seen:
            seen[h] = 0
            unique_headers.append(h)
        else:
            seen[h] += 1
            unique_headers.append(f"{h}.{seen[h]}")
    return unique_headers

def iter_xlsx_rows(file_path, sheet_name=None, skip_header=True):
    op = require_openpyxl()
    wb = op.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    start_row = 2 if skip_header else 1
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        yield ["" if c is None else str(c) for c in row]
    wb.close()

def is_blank(value):
    """Returns True if the value is empty or contains only whitespace."""
    if value is None:
        return True
    return str(value).strip() == ""

def guess_file_type(path):
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".txt"):
        return "txt"
    return "unknown"


# ---------- Advanced VLOOKUP Logic ----------

def build_reference_map(ref_path, ref_key_col, data_cols, ref_sep=None, ref_encoding=None,
                         allow_repeats=False, ref_sheet=None):
    filetype = guess_file_type(ref_path)
    ref_map = {}
    max_len_per_col = {c: 1 for c in data_cols}

    if filetype in ("csv", "txt"):
        headers, sep, enc = read_text_header(ref_path, sep=ref_sep, encoding=ref_encoding)
        try:
            key_idx = headers.index(ref_key_col)
        except ValueError:
            raise RuntimeError(f"Key column '{ref_key_col}' does not exist in the reference file.")
        data_idx = []
        for c in data_cols:
            try:
                data_idx.append(headers.index(c))
            except ValueError:
                raise RuntimeError(f"Column '{c}' does not exist in the reference file.")

        for row in iter_text_rows(ref_path, sep, enc, skip_header=True):
            key = row[key_idx] if key_idx < len(row) else ""
            if allow_repeats:
                d = ref_map.get(key)
                if d is None:
                    d = {c: [] for c in data_cols}
                    ref_map[key] = d
                for c, ix in zip(data_cols, data_idx):
                    val = row[ix] if ix < len(row) else ""
                    if not is_blank(val):
                        d[c].append(val)
                    if len(d[c]) > max_len_per_col[c]:
                        max_len_per_col[c] = len(d[c])
            else:
                if key in ref_map:
                    continue
                vals = {}
                for c, ix in zip(data_cols, data_idx):
                    vals[c] = row[ix] if ix < len(row) else ""
                ref_map[key] = vals

    elif filetype == "xlsx":
        headers = read_xlsx_header(ref_path, sheet_name=ref_sheet)
        try:
            key_idx = headers.index(ref_key_col)
        except ValueError:
            raise RuntimeError(f"Key column '{ref_key_col}' does not exist in the reference file.")
        data_idx = []
        for c in data_cols:
            try:
                data_idx.append(headers.index(c))
            except ValueError:
                raise RuntimeError(f"Column '{c}' does not exist in the reference file.")

        for row in iter_xlsx_rows(ref_path, sheet_name=ref_sheet, skip_header=True):
            key = row[key_idx] if key_idx < len(row) else ""
            if allow_repeats:
                d = ref_map.get(key)
                if d is None:
                    d = {c: [] for c in data_cols}
                    ref_map[key] = d
                for c, ix in zip(data_cols, data_idx):
                    val = row[ix] if ix < len(row) else ""
                    if not is_blank(val):
                        d[c].append(val)
                    if len(d[c]) > max_len_per_col[c]:
                        max_len_per_col[c] = len(d[c])
            else:
                if key in ref_map:
                    continue
                vals = {}
                for c, ix in zip(data_cols, data_idx):
                    vals[c] = row[ix] if ix < len(row) else ""
                ref_map[key] = vals
    else:
        raise RuntimeError("Unsupported reference file format. Use .txt, .csv, or .xlsx.")
    return ref_map, max_len_per_col

def stream_join_and_write(princ_path, ref_map, princ_key_col, selected_princ_cols, data_cols,
                           only_matches, allow_repeats, max_len_per_col, out_format, out_folder,
                           out_base_name, split_files, max_lines, princ_sep=None,
                           princ_encoding=None, princ_sheet=None):
    filetype = guess_file_type(princ_path)
    if filetype in ("csv", "txt"):
        headers, sep, enc = read_text_header(princ_path, sep=princ_sep, encoding=princ_encoding)
        rows_iter = iter_text_rows(princ_path, sep, enc, skip_header=True)
        out_sep = TEXT_SEP_DEFAULT_TXT if out_format == ".txt" else TEXT_SEP_DEFAULT_CSV
    elif filetype == "xlsx":
        headers = read_xlsx_header(princ_path, sheet_name=princ_sheet)
        rows_iter = iter_xlsx_rows(princ_path, sheet_name=princ_sheet, skip_header=True)
        out_sep = TEXT_SEP_DEFAULT_TXT if out_format == ".txt" else TEXT_SEP_DEFAULT_CSV
    else:
        raise RuntimeError("Unsupported main file format. Use .txt, .csv, or .xlsx.")

    try:
        princ_key_idx = headers.index(princ_key_col)
    except ValueError:
        raise RuntimeError(f"Key column '{princ_key_col}' does not exist in the main file.")

    if not selected_princ_cols or len(selected_princ_cols) == 0:
        princ_keep_idx = list(range(len(headers)))
        princ_keep_names = headers[:]
    else:
        princ_keep_idx = []
        princ_keep_names = []
        for c in selected_princ_cols:
            try:
                ix = headers.index(c)
            except ValueError:
                raise RuntimeError(f"Column '{c}' does not exist in the main file.")
            princ_keep_idx.append(ix)
            princ_keep_names.append(c)

    imported_col_names = []
    if allow_repeats:
        for col in data_cols:
            m = max_len_per_col.get(col, 1)
            for i in range(1, m + 1):
                imported_col_names.append(col if i == 1 else f"{col}_{i}")
    else:
        imported_col_names.extend(data_cols)

    final_headers = princ_keep_names + imported_col_names
    imported_total = len(imported_col_names)

    def open_new_writer(file_index):
        suffix = f"_part{file_index:03d}" if split_files else ""
        out_name = f"{out_base_name}{suffix}{out_format}"
        out_path = os.path.join(out_folder, out_name)
        if out_format in (".csv", ".txt"):
            f = open(out_path, "w", encoding="utf-8", newline="")
            writer = csv.writer(f, delimiter=out_sep)
            writer.writerow(final_headers)
            return ("text", out_path, f, writer)
        elif out_format == ".xlsx":
            xw = require_xlsxwriter()
            wb = xw.Workbook(out_path, {"strings_to_numbers": False, "strings_to_formulas": False})
            ws = wb.add_worksheet("Result")
            text_fmt = wb.add_format({"num_format": "@"})
            for col_idx, val in enumerate(final_headers):
                ws.write_string(0, col_idx, str(val), text_fmt)
            return ("xlsx", out_path, wb, ws, text_fmt, 1)
        else:
            raise RuntimeError("Unknown output format.")

    def close_writer(handle):
        if handle is None:
            return
        kind = handle[0]
        if kind == "text":
            _, _, f, _ = handle
            f.close()
        elif kind == "xlsx":
            _, _, wb, _, _, _ = handle
            wb.close()

    os.makedirs(out_folder, exist_ok=True)
    file_count = 1
    written_in_current = 0
    handle = open_new_writer(file_count)

    def write_row(values):
        nonlocal handle, file_count, written_in_current
        if split_files and written_in_current >= max_lines:
            close_writer(handle)
            file_count += 1
            written_in_current = 0
            handle = open_new_writer(file_count)
        kind = handle[0]
        if kind == "text":
            _, _, _, writer = handle
            writer.writerow([str(v) if v is not None else "" for v in values])
        elif kind == "xlsx":
            _, _, _, ws, text_fmt, next_row = handle
            for cidx, v in enumerate(values):
                ws.write_string(next_row, cidx, "" if v is None else str(v), text_fmt)
            handle = (handle[0], handle[1], handle[2], handle[3], handle[4], next_row + 1)
        written_in_current += 1
        return handle

    total_written = 0
    for row in rows_iter:
        key = row[princ_key_idx] if princ_key_idx < len(row) else ""
        match_vals = ref_map.get(key)

        if match_vals is None:
            if only_matches:
                continue
            importeds = [""] * imported_total
        else:
            importeds = []
            if allow_repeats:
                for col in data_cols:
                    lst = [v for v in match_vals.get(col, []) if not is_blank(v)]
                    m = max_len_per_col.get(col, 1)
                    if len(lst) < m:
                        lst += [""] * (m - len(lst))
                    else:
                        lst = lst[:m]
                    importeds.extend(lst)
            else:
                for col in data_cols:
                    importeds.append(match_vals.get(col, ""))

        out_base = [row[ix] if ix < len(row) else "" for ix in princ_keep_idx]
        out_row = out_base + importeds
        handle = write_row(out_row)
        total_written += 1

    close_writer(handle)
    return total_written, file_count


# ---------- UI (CustomTkinter) ----------

class MultiSelectDialog(ctk.CTkToplevel):
    def __init__(self, parent, title, options, preselected=None):
        super().__init__(parent)
        self.title(title)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.geometry("+%d+%d" % (parent.winfo_rootx() + 80, parent.winfo_rooty() + 80))
        self.minsize(350, 400)

        ctk.CTkLabel(self, text=title, font=ctk.CTkFont(weight="bold")).pack(padx=15, pady=(15, 5), anchor="w")

        self.scroll = ctk.CTkScrollableFrame(self)
        self.scroll.pack(fill="both", expand=True, padx=15, pady=5)

        self.vars = []
        pre = set(preselected or [])
        for opt in options:
            var = ctk.BooleanVar(value=(opt in pre))
            chk = ctk.CTkCheckBox(self.scroll, text=opt, variable=var)
            chk.pack(anchor="w", pady=5)
            self.vars.append((opt, var))

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(pady=15)
        ctk.CTkButton(btns, text="OK", command=self.on_ok, width=100).pack(side="left", padx=10)
        ctk.CTkButton(btns, text="Cancel", command=self.on_cancel, width=100, fg_color="gray", hover_color="#555555").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self.on_cancel)

    def on_ok(self):
        selected = [name for (name, var) in self.vars if var.get()]
        self.result = selected
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()

class ChooseSheetDialog(ctk.CTkToplevel):
    def __init__(self, parent, sheets, current=None, title="Choose Sheet"):
        super().__init__(parent)
        self.title(title)
        self.transient(parent)
        self.grab_set()
        self.resizable(False, False)
        self.result = None

        ctk.CTkLabel(self, text="Select a sheet:", font=ctk.CTkFont(weight="bold")).pack(padx=20, pady=(20, 10))

        self.var = ctk.StringVar(value=current if current in sheets else (sheets[0] if sheets else ""))
        self.cb = ctk.CTkComboBox(self, values=sheets, variable=self.var, state="readonly", width=250)
        self.cb.pack(padx=20, pady=10)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(pady=(15, 20))
        ctk.CTkButton(btns, text="OK", command=self.on_ok, width=100).pack(side="left", padx=10)
        ctk.CTkButton(btns, text="Cancel", command=self.on_cancel, width=100, fg_color="gray", hover_color="#555555").pack(side="left", padx=10)

        self.update_idletasks()
        _center_window(self, parent)

    def on_ok(self):
        self.result = self.var.get()
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()

class ChooseSourceDialog(ctk.CTkToplevel):
    def __init__(self, parent, has_main: bool, has_ref: bool):
        super().__init__(parent)
        self.title("Choose Source")
        self.transient(parent)
        self.grab_set()
        self.resizable(False, False)
        self.result = None

        ctk.CTkLabel(self, text="Which file do you want to export clean data from?", font=ctk.CTkFont(weight="bold")).pack(padx=20, pady=(20, 15))
        self.var = ctk.StringVar(value=("main" if has_main else "reference"))

        rb_frame = ctk.CTkFrame(self, fg_color="transparent")
        rb_frame.pack(padx=20, pady=(0, 15))

        self.rb_main = ctk.CTkRadioButton(rb_frame, text="Main File", variable=self.var, value="main", state=("normal" if has_main else "disabled"))
        self.rb_ref = ctk.CTkRadioButton(rb_frame, text="Reference File", variable=self.var, value="reference", state=("normal" if has_ref else "disabled"))
        self.rb_main.pack(anchor="w", pady=5)
        self.rb_ref.pack(anchor="w", pady=5)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(pady=(5, 20))
        ctk.CTkButton(btns, text="OK", command=self.on_ok, width=100).pack(side="left", padx=10)
        ctk.CTkButton(btns, text="Cancel", command=self.on_cancel, width=100, fg_color="gray", hover_color="#555555").pack(side="left", padx=10)

        self.update_idletasks()
        _center_window(self, parent)

    def on_ok(self):
        self.result = self.var.get()
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.icon_setup()

        self.geometry("1100x800")
        self.minsize(900, 600)
        try:
            self.state("zoomed")
        except Exception:
            pass

        # State vars
        self.main_path = ctk.StringVar(value="")
        self.ref_path = ctk.StringVar(value="")
        self.main_sep = ctk.StringVar(value="")
        self.ref_sep = ctk.StringVar(value="")
        self.main_encoding = ctk.StringVar(value="")
        self.ref_encoding = ctk.StringVar(value="")
        self.main_headers = []
        self.ref_headers = []

        self.main_key = ctk.StringVar(value="")
        self.ref_key = ctk.StringVar(value="")

        self.data_cols_selected = []
        self.final_cols_selected = []
        self.allow_repeats = ctk.BooleanVar(value=False)
        self.only_matches = ctk.BooleanVar(value=False)
        self.split_files = ctk.BooleanVar(value=False)
        self.max_lines_str = ctk.StringVar(value="100000")

        self.save_folder = ctk.StringVar(value=os.path.expanduser("~/Desktop"))
        self.save_format = ctk.StringVar(value=".csv")

        self.main_sheets = []
        self.ref_sheets = []
        self.main_sheet = ctk.StringVar(value="")
        self.ref_sheet = ctk.StringVar(value="")

        self.history = []
        self.history_lock = threading.Lock()

        self.setup_styles()
        self.build_ui()
        self.after(5000, self.history_watchdog)

    def icon_setup(self):
        try:
            base_dirs = [
                getattr(sys, "_MEIPASS", ""),
                os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")),
                (os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else ""),
            ]
            candidates = []
            for b in base_dirs:
                if b:
                    candidates.append(os.path.join(b, "assets", "icon.ico"))
            if getattr(sys, "frozen", False):
                candidates.append(os.path.join(os.path.dirname(sys.executable), "icon.ico"))

            for p in candidates:
                if p and os.path.exists(p):
                    if os.name == "nt":
                        self.iconbitmap(p)
                    break
        except Exception:
            pass

    def setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("default")

        bg_color = "#2b2b2b"
        fg_color = "white"
        sel_color = "#1f538d"
        head_bg = "#343638"

        style.configure("Treeview",
            background=bg_color,
            foreground=fg_color,
            rowheight=25,
            fieldbackground=bg_color,
            bordercolor=head_bg,
            borderwidth=0)
        style.map('Treeview', background=[('selected', sel_color)])
        style.configure("Treeview.Heading",
            background=head_bg,
            foreground=fg_color,
            relief="flat")
        style.map("Treeview.Heading", background=[('active', '#565b5e')])

    def build_ui(self):
        self.main_container = ctk.CTkFrame(self, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True)

        self.scroll_frame = ctk.CTkScrollableFrame(self.main_container, fg_color="transparent")
        self.scroll_frame.pack(side="top", fill="both", expand=True, padx=10, pady=10)

        def create_section(parent, title):
            wrapper = ctk.CTkFrame(parent)
            wrapper.pack(fill="x", pady=(0, 15), padx=5)
            lbl = ctk.CTkLabel(wrapper, text=title, font=ctk.CTkFont(size=14, weight="bold"))
            lbl.pack(anchor="w", padx=15, pady=(10, 0))
            inner = ctk.CTkFrame(wrapper, fg_color="transparent")
            inner.pack(fill="x", padx=15, pady=10)
            return inner

        # 1. Files Section
        files_group = create_section(self.scroll_frame, "Files")

        # Main file
        row1 = ctk.CTkFrame(files_group, fg_color="transparent")
        row1.pack(fill="x", pady=5)
        ctk.CTkLabel(row1, text="Load Main File:", width=150, anchor="w").pack(side="left")
        ctk.CTkEntry(row1, textvariable=self.main_path, width=400, state="disabled").pack(side="left", padx=10)
        ctk.CTkButton(row1, text="Browse", command=self.select_main_file, width=100).pack(side="left", padx=5)
        self.btn_main_sheet = ctk.CTkButton(row1, text="Choose Sheet", command=self.choose_main_sheet, state="disabled", width=100)
        self.btn_main_sheet.pack(side="left", padx=5)
        ctk.CTkLabel(row1, text="Separator:").pack(side="left", padx=(20, 5))
        self.main_sep_entry = ctk.CTkEntry(row1, textvariable=self.main_sep, width=40)
        self.main_sep_entry.pack(side="left")

        # Reference file
        row2 = ctk.CTkFrame(files_group, fg_color="transparent")
        row2.pack(fill="x", pady=5)
        ctk.CTkLabel(row2, text="Load Reference File:", width=150, anchor="w").pack(side="left")
        ctk.CTkEntry(row2, textvariable=self.ref_path, width=400, state="disabled").pack(side="left", padx=10)
        ctk.CTkButton(row2, text="Browse", command=self.select_reference_file, width=100).pack(side="left", padx=5)
        self.btn_ref_sheet = ctk.CTkButton(row2, text="Choose Sheet", command=self.choose_reference_sheet, state="disabled", width=100)
        self.btn_ref_sheet.pack(side="left", padx=5)
        ctk.CTkLabel(row2, text="Separator:").pack(side="left", padx=(20, 5))
        self.ref_sep_entry = ctk.CTkEntry(row2, textvariable=self.ref_sep, width=40)
        self.ref_sep_entry.pack(side="left")

        # 2. Column Configuration Section
        cols_group = create_section(self.scroll_frame, "Column Configuration")

        row3 = ctk.CTkFrame(cols_group, fg_color="transparent")
        row3.pack(fill="x", pady=5)
        ctk.CTkLabel(row3, text="Main File Key Column:", width=200, anchor="w").pack(side="left")
        self.main_key_combo = ctk.CTkComboBox(row3, values=[], variable=self.main_key, state="readonly", width=300)
        self.main_key_combo.pack(side="left", padx=10)

        row4 = ctk.CTkFrame(cols_group, fg_color="transparent")
        row4.pack(fill="x", pady=5)
        ctk.CTkLabel(row4, text="Reference File Key Column:", width=200, anchor="w").pack(side="left")
        self.ref_key_combo = ctk.CTkComboBox(row4, values=[], variable=self.ref_key, state="readonly", width=300)
        self.ref_key_combo.pack(side="left", padx=10)

        row5 = ctk.CTkFrame(cols_group, fg_color="transparent")
        row5.pack(fill="x", pady=5)
        ctk.CTkLabel(row5, text="Columns to Import:", width=200, anchor="w").pack(side="left")
        self.data_cols_label = ctk.CTkLabel(row5, text="[nothing selected]", text_color="gray")
        self.data_cols_label.pack(side="left", padx=10)
        ctk.CTkButton(row5, text="Select", command=self.choose_data_cols, width=100).pack(side="left", padx=10)

        row6 = ctk.CTkFrame(cols_group, fg_color="transparent")
        row6.pack(fill="x", pady=5)
        self.chk_repeats = ctk.CTkCheckBox(row6, text="Bring repeated values (all occurrences)", variable=self.allow_repeats)
        self.chk_repeats.pack(side="left", padx=(0, 20))
        self.chk_matches = ctk.CTkCheckBox(row6, text="Matched rows only", variable=self.only_matches)
        self.chk_matches.pack(side="left")

        row8 = ctk.CTkFrame(cols_group, fg_color="transparent")
        row8.pack(fill="x", pady=5)
        ctk.CTkLabel(row8, text="Output columns from Main File:").pack(side="left")
        self.final_cols_label = ctk.CTkLabel(row8, text="[all by default]", text_color="gray")
        self.final_cols_label.pack(side="left", padx=10)
        ctk.CTkButton(row8, text="Select", command=self.choose_final_cols, width=100).pack(side="left", padx=10)

        # 3. Output & Save Section
        out_group = create_section(self.scroll_frame, "Output & Save")

        row9 = ctk.CTkFrame(out_group, fg_color="transparent")
        row9.pack(fill="x", pady=5)
        ctk.CTkCheckBox(row9, text="Split into multiple files", variable=self.split_files, command=self.on_toggle_split).pack(side="left")
        ctk.CTkLabel(row9, text="Row limit:").pack(side="left", padx=(30, 10))
        self.max_lines_entry = ctk.CTkEntry(row9, textvariable=self.max_lines_str, width=120)
        self.max_lines_entry.pack(side="left")
        ctk.CTkLabel(row9, text="(positive integer, multiple of 100,000)", text_color="gray").pack(side="left", padx=10)

        row10 = ctk.CTkFrame(out_group, fg_color="transparent")
        row10.pack(fill="x", pady=10)
        ctk.CTkLabel(row10, text="Save folder:", width=120, anchor="w").pack(side="left")
        ctk.CTkEntry(row10, textvariable=self.save_folder, width=400, state="disabled").pack(side="left", padx=10)
        ctk.CTkButton(row10, text="Browse", command=self.select_save_folder, width=100).pack(side="left", padx=5)

        row11 = ctk.CTkFrame(out_group, fg_color="transparent")
        row11.pack(fill="x", pady=5)
        ctk.CTkLabel(row11, text="Output format:", width=120, anchor="w").pack(side="left")
        self.format_combo = ctk.CTkComboBox(row11, values=[".txt", ".csv", ".xlsx"], variable=self.save_format, state="readonly", width=100)
        self.format_combo.pack(side="left", padx=10)

        # 4. Preview Section
        prev_wrapper = ctk.CTkFrame(self.scroll_frame)
        prev_wrapper.pack(fill="x", pady=(0, 15), padx=5)
        ctk.CTkLabel(prev_wrapper, text="Preview (first 5 rows)", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))

        tree_frame = ctk.CTkFrame(prev_wrapper, fg_color="transparent")
        tree_frame.pack(fill="x", padx=15, pady=(0, 15))

        self.preview_tree = ttk.Treeview(tree_frame, columns=[], show="headings", height=8)
        self.preview_tree.pack(fill="x", expand=True)

        # 5. History Section
        hist_wrapper = ctk.CTkFrame(self.scroll_frame)
        hist_wrapper.pack(fill="x", pady=(0, 15), padx=5)
        ctk.CTkLabel(hist_wrapper, text="Saved Files History (today)", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))

        hist_inner = ctk.CTkFrame(hist_wrapper, fg_color="transparent")
        hist_inner.pack(fill="x", padx=15, pady=(0, 15))

        self.history_list = tk.Listbox(hist_inner, height=5, bg="#2b2b2b", fg="white",
            selectbackground="#1f538d", highlightthickness=0, bd=0, font=("Segoe UI", 10))
        self.history_list.pack(fill="x", expand=True, pady=(0, 10))

        btn_hist_row = ctk.CTkFrame(hist_inner, fg_color="transparent")
        btn_hist_row.pack(fill="x")
        ctk.CTkButton(btn_hist_row, text="Open Folder", command=self.open_selected_history_folder, width=150).pack(side="left")
        ctk.CTkButton(btn_hist_row, text="Refresh", command=self.refresh_history_view, width=150, fg_color="#444444", hover_color="#555555").pack(side="left", padx=10)

        # Fixed action bar at bottom
        button_bar = ctk.CTkFrame(self.main_container, height=60, corner_radius=0)
        button_bar.pack(side="bottom", fill="x")

        ctk.CTkButton(button_bar, text="Preview", command=self.on_preview, width=150, fg_color="#2c7a3f", hover_color="#399e52").pack(side="left", padx=20, pady=15)
        ctk.CTkButton(button_bar, text="Process", command=self.on_process, width=150, fg_color="#b85f00", hover_color="#d66e00").pack(side="left", padx=10, pady=15)
        ctk.CTkButton(button_bar, text="Reset", command=self.on_reset, width=100, fg_color="#7a1f1f", hover_color="#9e2a2a").pack(side="left", padx=10, pady=15)
        ctk.CTkButton(button_bar, text="Export Clean Data", command=self.on_export_clean, width=200).pack(side="right", padx=20, pady=15)

        self.on_toggle_split()

    # ---------- UI Methods / Logic ----------

    def open_selected_history_folder(self):
        sel = self.history_list.curselection()
        if not sel:
            messagebox.showwarning(APP_NAME, "Select an entry from the history list.")
            return

        path = self.history_list.get(sel[0])
        import subprocess

        folder = path if os.path.isdir(path) else (os.path.dirname(path) or ".")
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder], check=True)
            else:
                subprocess.run(["xdg-open", folder], check=True)
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Cannot open folder:\n{folder}\n\n{e}")

    def refresh_history_view(self):
        today = datetime.date.today()
        items = []
        with self.history_lock:
            new_hist = []
            for p, dt in self.history:
                if dt.date() == today and os.path.exists(p):
                    new_hist.append((p, dt))
            self.history = new_hist[:]
            items = [p for (p, _) in self.history]
        self.history_list.delete(0, "end")
        for p in items:
            self.history_list.insert("end", p)

    def _load_file_metadata(self, path, kind):
        ftype = guess_file_type(path)
        if ftype in ("csv", "txt"):
            headers, sep, enc = read_text_header(path)
            return {"path": path, "ftype": ftype, "headers": headers, "sep": sep or "", "enc": enc or "", "sheets": [], "sheet": ""}
        elif ftype == "xlsx":
            sheets = get_xlsx_sheet_names(path)
            chosen = sheets[0] if sheets else ""
            headers = read_xlsx_header(path, sheet_name=chosen)
            return {"path": path, "ftype": ftype, "headers": headers, "sep": "", "enc": "", "sheets": sheets, "sheet": chosen}
        else:
            raise RuntimeError("Unsupported format. Use .txt, .csv, or .xlsx.")

    def _load_headers_for_sheet(self, path, sheet):
        return read_xlsx_header(path, sheet_name=sheet)

    def select_main_file(self):
        path = filedialog.askopenfilename(
            title="Select Main File",
            filetypes=[("All supported", "*.txt;*.csv;*.xlsx"), ("Text", "*.txt"), ("CSV", "*.csv"), ("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not path:
            return

        def cb(result, error):
            if error:
                messagebox.showerror(APP_NAME, f"Error reading Main File:\n{error}")
                return
            self.main_path.set(result["path"])
            self.main_headers = result["headers"]
            self.main_sep.set(result["sep"])
            self.main_encoding.set(result["enc"])

            self.main_key_combo.configure(values=self.main_headers)
            self.main_key.set("")

            self.main_sheets = result["sheets"]
            self.main_sheet.set(result["sheet"])
            is_xlsx = (result["ftype"] == "xlsx")
            self.btn_main_sheet.configure(state=("normal" if is_xlsx and len(self.main_sheets) > 1 else "disabled"))
            self.main_sep_entry.configure(state=("disabled" if is_xlsx else "normal"))

            self.final_cols_selected = []
            self.final_cols_label.configure(text="[all by default]")

        run_with_busy(self, self._load_file_metadata, cb, path, "main", message="Loading file… please wait.")

    def select_reference_file(self):
        path = filedialog.askopenfilename(
            title="Select Reference File",
            filetypes=[("All supported", "*.txt;*.csv;*.xlsx"), ("Text", "*.txt"), ("CSV", "*.csv"), ("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not path:
            return

        def cb(result, error):
            if error:
                messagebox.showerror(APP_NAME, f"Error reading Reference File:\n{error}")
                return
            self.ref_path.set(result["path"])
            self.ref_headers = result["headers"]
            self.ref_sep.set(result["sep"])
            self.ref_encoding.set(result["enc"])

            self.ref_key_combo.configure(values=self.ref_headers)
            self.ref_key.set("")

            self.ref_sheets = result["sheets"]
            self.ref_sheet.set(result["sheet"])
            is_xlsx = (result["ftype"] == "xlsx")
            self.btn_ref_sheet.configure(state=("normal" if is_xlsx and len(self.ref_sheets) > 1 else "disabled"))
            self.ref_sep_entry.configure(state=("disabled" if is_xlsx else "normal"))

            self.data_cols_selected = []
            self.data_cols_label.configure(text="[nothing selected]")

        run_with_busy(self, self._load_file_metadata, cb, path, "ref", message="Loading file… please wait.")

    def choose_main_sheet(self):
        if not self.main_path.get():
            messagebox.showinfo(APP_NAME, "Load the Main File first.")
            return
        if not self.main_sheets or len(self.main_sheets) <= 1:
            messagebox.showinfo(APP_NAME, "This file does not have multiple sheets.")
            return
        dlg = ChooseSheetDialog(self, self.main_sheets, current=self.main_sheet.get(), title="Choose Sheet — Main File")
        self.wait_window(dlg)
        if dlg.result and dlg.result != self.main_sheet.get():
            def cb(headers, error):
                if error:
                    messagebox.showerror(APP_NAME, f"Error reading sheet headers:\n{error}")
                    return
                self.main_sheet.set(dlg.result)
                self.main_headers = headers
                self.main_key_combo.configure(values=self.main_headers)
                self.main_key.set("")
                self.final_cols_selected = []
                self.final_cols_label.configure(text="[all by default]")
            run_with_busy(self, self._load_headers_for_sheet, cb, self.main_path.get(), dlg.result, message="Loading file… please wait.")

    def choose_reference_sheet(self):
        if not self.ref_path.get():
            messagebox.showinfo(APP_NAME, "Load the Reference File first.")
            return
        if not self.ref_sheets or len(self.ref_sheets) <= 1:
            messagebox.showinfo(APP_NAME, "This file does not have multiple sheets.")
            return
        dlg = ChooseSheetDialog(self, self.ref_sheets, current=self.ref_sheet.get(), title="Choose Sheet — Reference File")
        self.wait_window(dlg)
        if dlg.result and dlg.result != self.ref_sheet.get():
            def cb(headers, error):
                if error:
                    messagebox.showerror(APP_NAME, f"Error reading sheet headers:\n{error}")
                    return
                self.ref_sheet.set(dlg.result)
                self.ref_headers = headers
                self.ref_key_combo.configure(values=self.ref_headers)
                self.ref_key.set("")
                self.data_cols_selected = []
                self.data_cols_label.configure(text="[nothing selected]")
            run_with_busy(self, self._load_headers_for_sheet, cb, self.ref_path.get(), dlg.result, message="Loading file… please wait.")

    def choose_data_cols(self):
        if not self.ref_headers:
            messagebox.showinfo(APP_NAME, "Load the Reference File first.")
            return
        dlg = MultiSelectDialog(self, "Select columns to import", options=self.ref_headers, preselected=self.data_cols_selected)
        self.wait_window(dlg)
        if dlg.result is not None:
            self.data_cols_selected = dlg.result
            text = (", ".join(self.data_cols_selected) if self.data_cols_selected else "[nothing selected]")
            self.data_cols_label.configure(text=text)

    def choose_final_cols(self):
        if not self.main_headers:
            messagebox.showinfo(APP_NAME, "Load the Main File first.")
            return
        dlg = MultiSelectDialog(self, "Choose output columns (all by default)", options=self.main_headers, preselected=self.final_cols_selected)
        self.wait_window(dlg)
        if dlg.result is not None:
            self.final_cols_selected = dlg.result
            text = (", ".join(self.final_cols_selected) if self.final_cols_selected else "[all by default]")
            self.final_cols_label.configure(text=text)

    def on_toggle_split(self):
        if self.split_files.get():
            self.max_lines_entry.configure(state="normal")
        else:
            self.max_lines_entry.configure(state="disabled")

    def select_save_folder(self):
        d = filedialog.askdirectory(title="Select save folder", initialdir=self.save_folder.get() or os.path.expanduser("~/Desktop"))
        if d:
            self.save_folder.set(d)

    def ensure_valid_config(self):
        if not self.main_path.get(): raise RuntimeError("Select the Main File.")
        if not self.ref_path.get(): raise RuntimeError("Select the Reference File.")
        if not self.main_key.get(): raise RuntimeError("Select the 'Main File Key Column'.")
        if not self.ref_key.get(): raise RuntimeError("Select the 'Reference File Key Column'.")
        if not self.data_cols_selected: raise RuntimeError("Select at least one column under 'Columns to Import'.")

        if self.split_files.get():
            try:
                n = int(self.max_lines_str.get())
                if n <= 0 or n % 100000 != 0: raise ValueError()
            except Exception:
                raise RuntimeError("'Row limit' must be a positive integer, multiple of 100,000.")

        if not self.save_folder.get(): raise RuntimeError("Select a 'Save folder'.")
        fmt = self.save_format.get()
        if fmt not in [".txt", ".csv", ".xlsx"]: raise RuntimeError("Select an 'Output format' (.txt, .csv, or .xlsx).")

        for path, sep_var, label in [(self.main_path.get(), self.main_sep, "main"), (self.ref_path.get(), self.ref_sep, "reference")]:
            ftype = guess_file_type(path)
            if ftype in ("csv", "txt"):
                sep = sep_var.get()
                if not sep: raise RuntimeError(f"Separator for the {label} file is not set.")
                if len(sep) > 1 and sep not in ["\\t", "\\x09"]:
                    messagebox.showwarning(APP_NAME, f"Separator for the {label} file has more than one character. Make sure it is correct.")
        return True

    def on_preview(self):
        try: self.ensure_valid_config()
        except Exception as e:
            messagebox.showerror(APP_NAME, str(e))
            return

        try:
            allow_repeats = self.allow_repeats.get()
            ref_map, max_len_per_col = build_reference_map(
                self.ref_path.get(), self.ref_key.get(), self.data_cols_selected,
                ref_sep=(self.ref_sep.get() or None), ref_encoding=(self.ref_encoding.get() or None),
                allow_repeats=allow_repeats, ref_sheet=(self.ref_sheet.get() or None),
            )

            main_type = guess_file_type(self.main_path.get())
            if main_type in ("csv", "txt"):
                headers, sep, enc = read_text_header(self.main_path.get(), sep=(self.main_sep.get() or None), encoding=(self.main_encoding.get() or None))
                rows = []
                it = iter_text_rows(self.main_path.get(), sep, enc, skip_header=True)
                for _ in range(5):
                    try: rows.append(next(it))
                    except StopIteration: break
            else:
                headers = read_xlsx_header(self.main_path.get(), sheet_name=(self.main_sheet.get() or None))
                rows = []
                it = iter_xlsx_rows(self.main_path.get(), sheet_name=(self.main_sheet.get() or None), skip_header=True)
                for _ in range(5):
                    try: rows.append(next(it))
                    except StopIteration: break

            if not self.final_cols_selected:
                princ_keep_idx = list(range(len(headers)))
                princ_keep_names = headers[:]
            else:
                princ_keep_idx = []
                princ_keep_names = []
                for c in self.final_cols_selected:
                    ix = headers.index(c)
                    princ_keep_idx.append(ix)
                    princ_keep_names.append(c)

            imported = []
            if allow_repeats:
                for col in self.data_cols_selected:
                    m = max_len_per_col.get(col, 1)
                    for i in range(1, m + 1):
                        imported.append(col if i == 1 else f"{col}_{i}")
            else:
                imported.extend(self.data_cols_selected)

            cols_final = princ_keep_names + imported

            self.preview_tree.delete(*self.preview_tree.get_children())
            self.preview_tree["columns"] = list(range(len(cols_final)))
            for i, name in enumerate(cols_final):
                self.preview_tree.heading(i, text=name)
                self.preview_tree.column(i, width=160, anchor="w")

            pkey_idx = headers.index(self.main_key.get())
            for r in rows:
                key = r[pkey_idx] if pkey_idx < len(r) else ""
                match_vals = ref_map.get(key)
                if match_vals is None:
                    if self.only_matches.get(): continue
                    imp_vals = []
                    if allow_repeats:
                        for col in self.data_cols_selected:
                            m = max_len_per_col.get(col, 1)
                            imp_vals.extend([""] * m)
                    else:
                        imp_vals.extend([""] * len(self.data_cols_selected))
                else:
                    imp_vals = []
                    if allow_repeats:
                        for col in self.data_cols_selected:
                            lst = [v for v in match_vals.get(col, []) if not is_blank(v)]
                            m = max_len_per_col.get(col, 1)
                            if len(lst) < m:
                                lst += [""] * (m - len(lst))
                            else:
                                lst = lst[:m]
                            imp_vals.extend(lst)
                    else:
                        for col in self.data_cols_selected:
                            imp_vals.append(match_vals.get(col, ""))

                base = [r[ix] if ix < len(r) else "" for ix in princ_keep_idx]
                row_out = base + imp_vals
                self.preview_tree.insert("", "end", values=row_out)
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Preview error:\n{e}")

    def on_reset(self):
        if not messagebox.askyesno(APP_NAME, "Are you sure you want to reset all selections?"):
            return

        self.main_path.set("")
        self.ref_path.set("")
        self.main_sep.set("")
        self.ref_sep.set("")
        self.main_encoding.set("")
        self.ref_encoding.set("")
        self.main_headers = []
        self.ref_headers = []
        self.main_sheets = []
        self.ref_sheets = []
        self.main_sheet.set("")
        self.ref_sheet.set("")

        self.main_key.set("")
        self.ref_key.set("")
        self.data_cols_selected = []
        self.final_cols_selected = []

        self.allow_repeats.set(False)
        self.only_matches.set(False)
        self.split_files.set(False)
        self.max_lines_str.set("100000")

        self.save_format.set(".csv")
        self.save_folder.set(os.path.expanduser("~/Desktop"))

        self.main_key_combo.configure(values=[])
        self.ref_key_combo.configure(values=[])

        self.data_cols_label.configure(text="[nothing selected]")
        self.final_cols_label.configure(text="[all by default]")

        self.btn_main_sheet.configure(state="disabled")
        self.btn_ref_sheet.configure(state="disabled")

        self.main_sep_entry.configure(state="normal")
        self.ref_sep_entry.configure(state="normal")

        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = []

        self.on_toggle_split()

    def on_process(self):
        try: self.ensure_valid_config()
        except Exception as e:
            messagebox.showerror(APP_NAME, str(e))
            return
        run_with_busy(self, self._do_process, self._on_process_done, message="Processing… please wait.")

    def _do_process(self):
        allow_repeats = self.allow_repeats.get()
        fmt = self.save_format.get()
        if fmt == ".xlsx": require_xlsxwriter()

        ref_map, max_len_per_col = build_reference_map(
            self.ref_path.get(), self.ref_key.get(), self.data_cols_selected,
            ref_sep=(self.ref_sep.get() or None), ref_encoding=(self.ref_encoding.get() or None),
            allow_repeats=allow_repeats, ref_sheet=(self.ref_sheet.get() or None),
        )

        out_folder = self.save_folder.get()
        base_name = os.path.splitext(os.path.basename(self.main_path.get()))[0] + "_result"
        split_files = self.split_files.get()
        max_lines = int(self.max_lines_str.get()) if split_files else 0

        total_written, file_count = stream_join_and_write(
            princ_path=self.main_path.get(), ref_map=ref_map, princ_key_col=self.main_key.get(),
            selected_princ_cols=self.final_cols_selected, data_cols=self.data_cols_selected,
            only_matches=self.only_matches.get(), allow_repeats=allow_repeats, max_len_per_col=max_len_per_col,
            out_format=fmt, out_folder=out_folder, out_base_name=base_name,
            split_files=split_files, max_lines=max_lines, princ_sep=(self.main_sep.get() or None),
            princ_encoding=(self.main_encoding.get() or None), princ_sheet=(self.main_sheet.get() or None),
        )

        return {"total_written": total_written, "file_count": file_count, "out_folder": out_folder, "base_name": base_name, "fmt": fmt, "split_files": split_files}

    def _on_process_done(self, result, error):
        if error:
            messagebox.showerror(APP_NAME, str(error))
            return

        if result["split_files"]:
            for i in range(1, result["file_count"] + 1):
                out_name = f'{result["base_name"]}_part{i:03d}{result["fmt"]}'
                out_path = os.path.join(result["out_folder"], out_name)
                self.add_history(out_path)
        else:
            out_path = os.path.join(result["out_folder"], f'{result["base_name"]}{result["fmt"]}')
            self.add_history(out_path)

        messagebox.showinfo(APP_NAME, f'Processing complete.\nRows written: {result["total_written"]}\nFiles generated: {result["file_count"]}')

    def on_export_clean(self):
        has_main = bool(self.main_path.get())
        has_ref = bool(self.ref_path.get())
        if not has_main and not has_ref:
            messagebox.showinfo(APP_NAME, "No Main File or Reference File has been selected.")
            return

        dlg = ChooseSourceDialog(self, has_main, has_ref)
        self.wait_window(dlg)
        if not dlg.result:
            return

        run_with_busy(self, self._do_export_clean, self._on_export_clean_done, dlg.result, message="Exporting… please wait.")

    def _do_export_clean(self, source_choice: str):
        require_xlsxwriter()
        xlsxwriter = sys.modules.get("xlsxwriter")

        if source_choice == "main":
            path = self.main_path.get()
            chosen_sheet = self.main_sheet.get() or None
        else:
            path = self.ref_path.get()
            chosen_sheet = self.ref_sheet.get() or None

        if not path: raise RuntimeError("Source file path is not set.")

        ftype = guess_file_type(path)
        out_folder = self.save_folder.get() or os.path.expanduser("~/Desktop")
        os.makedirs(out_folder, exist_ok=True)
        base = os.path.splitext(os.path.basename(path))[0]
        out_path = os.path.join(out_folder, f"{base}_clean.xlsx")

        wb = xlsxwriter.Workbook(out_path, {"strings_to_numbers": False, "strings_to_formulas": False})
        ws = wb.add_worksheet("Data")
        text_fmt = wb.add_format({"num_format": "@"})

        row_idx = 0
        if ftype in ("csv", "txt"):
            headers, sep, enc = read_text_header(path)
            for cidx, h in enumerate(headers):
                ws.write_string(row_idx, cidx, h if h is not None else "", text_fmt)
            row_idx += 1
            for r in iter_text_rows(path, sep, enc, skip_header=True):
                for cidx, v in enumerate(r):
                    ws.write_string(row_idx, cidx, "" if v is None else str(v), text_fmt)
                row_idx += 1
        elif ftype == "xlsx":
            headers = read_xlsx_header(path, sheet_name=chosen_sheet)
            for cidx, h in enumerate(headers):
                ws.write_string(row_idx, cidx, h if h is not None else "", text_fmt)
            row_idx += 1
            for r in iter_xlsx_rows(path, sheet_name=chosen_sheet, skip_header=True):
                for cidx, v in enumerate(r):
                    ws.write_string(row_idx, cidx, "" if v is None else str(v), text_fmt)
                row_idx += 1
        else:
            wb.close()
            raise RuntimeError("Unsupported file format for clean export.")

        wb.close()
        return out_path

    def _on_export_clean_done(self, result, error):
        if error:
            messagebox.showerror(APP_NAME, f"Error during clean export:\n{error}")
            return
        self.add_history(result)
        messagebox.showinfo(APP_NAME, f"File generated successfully:\n{result}")

    def add_history(self, path):
        try:
            dt = datetime.datetime.now()
            with self.history_lock:
                self.history.append((path, dt))
            self.refresh_history_view()
        except Exception:
            pass

    def history_watchdog(self):
        self.refresh_history_view()
        self.after(10000, self.history_watchdog)

def main():
    app = App()
    try:
        app.update_idletasks()
        app.deiconify()
        app.lift()
        app.attributes("-topmost", True)
        app.after(200, lambda: app.attributes("-topmost", False))
    except Exception as e:
        print("UI placement error:", e)
    app.mainloop()

if __name__ == "__main__":
    main()
